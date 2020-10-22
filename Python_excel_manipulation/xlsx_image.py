from openpyxl import Workbook 
from openpyxl.drawing.image import Image
import datetime

wb = Workbook()
ws = wb.active
ws['A1'] = datetime.datetime(2020,10,23)
print(ws['A1'].number_format)

#Apply equation to cell
ws['A2'] = '=SUM(1,1)'

#merge cells
ws.merge_cells('A3:D3')

#unmerge cells
#ws.unmerge_cells('A3:D3')

#insert an image
ws['A7'] = "You see image below"
img = Image('logo.jpg')
ws.add_image(img, 'A9')

#create new sheet with name 'Dim'
ws2 = wb.create_sheet(title='Dim')

#fold
ws2.column_dimensions.group('A','D', hidden=True)
ws2.row_dimensions.group(1,10, hidden=True)

wb.save('style.xlsx')